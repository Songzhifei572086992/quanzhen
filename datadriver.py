import openpyxl
from key_word import Webkey
from weblog import Logger
from time import sleep

def data_dr(sheet_number):
    #操作工作簿
    excel=openpyxl.load_workbook(r'C:\Users\全诊医学\PycharmProjects\QZ_webui\exl.xlsx')
    # print(excel)
    #获取sheet：基于工作簿来获取sheet
    # names=excel.sheetnames
    #操作单元格cell：制定sheet页，再进行操作
    sheet=excel[sheet_number]
    #单元格写入 sheet.cell(row='行数'，column=’列数‘).value='值'
    #读取Excel内容
    Logger().get_log().info('开始读取测试用例')
    for value in sheet.values:
        kwargs={}
        kwargs['name']=value[2]
        kwargs['values']=value[3]
        kwargs['send_text']=value[4]
        kwargs['yq_text']=value[6]
        if type(value[0]) is int:
           if value[1] == 'open_browser':
                wk = Webkey(value[4])
           elif 'assert' in value[1]:
               Logger().get_log().info('开始执行关键字: {},操作描述：{}'.format(value[1],value[5]))
               resuit = getattr(wk, value[1])(**kwargs)
               if resuit :
                   sheet.cell(row=value[0]+1,column=8).value = 'pass'
               else:
                   sheet.cell(row=value[0] + 1, column=8).value = 'failed'
           else:
                Logger().get_log().info('开始执行关键字: {},操作描述：{}'.format(value[1], value[5]))
                sleep(2)
                getattr(wk,value[1])(**kwargs)
                sleep(2)
           excel.save(r'C:\Users\全诊医学\PycharmProjects\QZ_webui\exl.xlsx')
           excel.close()
# data_dr('Sheet3')